from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import DonorRecord

REQUIRED_HEADERS = ["NO", "ÜLKE", "İL", "BÖLGE", "AD", "SOYAD", "TEL"]


@dataclass
class ImportSummary:
    imported_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
    missing_required_count: int = 0
    invalid_phone_count: int = 0
    duplicate_in_file_count: int = 0
    duplicate_in_db_count: int = 0
    errors: list[str] | None = None

    def __post_init__(self) -> None:
        if self.errors is None:
            self.errors = []


def _normalize_phone_e164(raw_phone: str) -> str | None:
    raw = str(raw_phone or "").strip()
    if not raw:
        return None

    compact = re.sub(r"[\s\-\(\)\.]", "", raw)
    if compact.startswith("00"):
        compact = "+" + compact[2:]
    elif compact.startswith("+"):
        compact = "+" + re.sub(r"\D", "", compact[1:])
    else:
        compact = "+" + re.sub(r"\D", "", compact)

    if not re.fullmatch(r"^\+[1-9]\d{7,14}$", compact):
        return None
    return compact


def _upsert_row(
    db: Session,
    row: dict[str, str],
    batch_id: str,
    org_id: int,
    summary: ImportSummary,
    row_no: int,
    seen_in_file: set[tuple[str, str, str, str]],
) -> None:
    no = str(row.get("NO", "")).strip()
    country = str(row.get("ÜLKE", "")).strip()
    city = str(row.get("İL", "")).strip()
    region = str(row.get("BÖLGE", "")).strip()
    first = str(row.get("AD", "")).strip()
    last = str(row.get("SOYAD", "")).strip()
    phone_raw = str(row.get("TEL", "")).strip()

    missing = [key for key, value in {"NO": no, "ÜLKE": country, "İL": city, "BÖLGE": region, "AD": first, "SOYAD": last, "TEL": phone_raw}.items() if not value]
    if missing:
        summary.skipped_count += 1
        summary.missing_required_count += 1
        summary.errors.append(f"Row {row_no}: missing required field(s): {', '.join(missing)}")
        return

    phone_e164 = _normalize_phone_e164(phone_raw)
    if not phone_e164:
        summary.skipped_count += 1
        summary.invalid_phone_count += 1
        summary.errors.append(f"Row {row_no}: invalid phone for E.164 format: {phone_raw}")
        return

    dedupe_key = (no, first.upper(), last.upper(), phone_e164)
    if dedupe_key in seen_in_file:
        summary.skipped_count += 1
        summary.duplicate_in_file_count += 1
        summary.errors.append(f"Row {row_no}: duplicate row in import file")
        return
    seen_in_file.add(dedupe_key)

    phone_digits = re.sub(r"\D", "", phone_raw)

    existing = (
        db.query(DonorRecord)
        .filter(
            DonorRecord.org_id == org_id,
            DonorRecord.no == no,
            DonorRecord.first_name == first,
            DonorRecord.last_name == last,
            DonorRecord.phone.in_({phone_e164, phone_raw, phone_digits}),
        )
        .first()
    )

    if existing:
        existing.country = country
        existing.city = city
        existing.region = region
        existing.phone = phone_e164
        existing.source_batch_id = batch_id
        existing.org_id = org_id
        summary.updated_count += 1
        summary.duplicate_in_db_count += 1
        return

    db.add(
        DonorRecord(
            no=no,
            country=country,
            city=city,
            region=region,
            first_name=first,
            last_name=last,
            phone=phone_e164,
            source_batch_id=batch_id,
            org_id=org_id,
        )
    )
    summary.imported_count += 1


def import_file(db: Session, file_path: Path, batch_id: str, org_id: int = 1) -> ImportSummary:
    summary = ImportSummary()
    seen_in_file: set[tuple[str, str, str, str]] = set()

    if file_path.suffix.lower() == ".csv":
        with file_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            if any(h not in headers for h in REQUIRED_HEADERS):
                raise ValueError(f"CSV headers must include: {REQUIRED_HEADERS}")
            for i, row in enumerate(reader, start=2):
                _upsert_row(db, row, batch_id, org_id, summary, i, seen_in_file)
    elif file_path.suffix.lower() in {".xlsx", ".xlsm"}:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise ValueError("Excel file is empty")
        mapped_headers = [str(x).strip() if x is not None else "" for x in header]
        if any(h not in mapped_headers for h in REQUIRED_HEADERS):
            raise ValueError(f"Excel headers must include: {REQUIRED_HEADERS}")
        index_map = {h: mapped_headers.index(h) for h in REQUIRED_HEADERS}

        for row_no, row in enumerate(rows, start=2):
            row_dict = {}
            for key, idx in index_map.items():
                row_dict[key] = "" if idx >= len(row) or row[idx] is None else str(row[idx])
            _upsert_row(db, row_dict, batch_id, org_id, summary, row_no, seen_in_file)
    else:
        raise ValueError("Unsupported file type. Use .xlsx or .csv")

    db.commit()
    return summary
